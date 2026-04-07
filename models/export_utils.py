"""
export_utils.py
Export δεδομένων σε CSV και Excel (.xlsx).
"""

import csv
import os
from datetime import datetime

from database.db_manager import status_label
from models.crud import get_all_samples, get_all_batches, get_all_clients


def export_samples_csv(db_path: str, output_path: str, status_filter=None) -> int:
    """
    Εξάγει όλα τα δείγματα σε CSV.
    Επιστρέφει πλήθος εξαχθέντων εγγραφών.
    """
    samples = get_all_samples(db_path, status_filter)

    headers = [
        "Κωδικός Δείγματος", "Batch", "Πελάτης", "Ανάλυση",
        "Κατάσταση", "Αναλυτής", "Αποτέλεσμα", "Παρατηρήσεις Αποτελέσματος",
        "Ημ. Αποτελέσματος", "Αναμ. Ημέρες", "Ημ. Εισαγωγής", "Τελ. Ενημέρωση"
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        for s in samples:
            writer.writerow([
                s['sample_code'],
                s['batch_code'],
                s['client_name'],
                s['analysis_name'],
                status_label(s['status']),
                s.get('analyst_name') or "",
                s.get('result_value') or "",
                s.get('result_notes') or "",
                s.get('result_at') or "",
                s.get('expected_days') or "",
                s['created_at'][:10],
                s['updated_at'][:10],
            ])

    return len(samples)


def export_samples_excel(db_path: str, output_path: str, status_filter=None) -> int:
    """
    Εξάγει δείγματα σε Excel (.xlsx) με χρωματισμό κατάστασης.
    Απαιτεί: pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Απαιτείται η βιβλιοθήκη openpyxl.\nΕκτελέστε: pip install openpyxl")

    from database.db_manager import STATUS_COLORS

    samples = get_all_samples(db_path, status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Δείγματα"

    # ── Header ────────────────────────────────────────────────────────────
    headers = [
        "Κωδικός Δείγματος", "Batch", "Πελάτης", "Ανάλυση",
        "Κατάσταση", "Αναλυτής", "Αποτέλεσμα", "Παρατηρήσεις",
        "Ημ. Αποτελέσματος", "Αναμ. Ημέρες", "Ημ. Εισαγωγής"
    ]

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 28

    # ── Data ──────────────────────────────────────────────────────────────
    STATUS_HEX = {
        "RECEIVED":    "3498DB",
        "IN_ANALYSIS": "E67E22",
        "ANALYZED":    "9B59B6",
        "PENDING":     "E74C3C",
        "COMPLETED":   "27AE60",
        "SHIPPED":     "95A5A6",
    }

    for row_idx, s in enumerate(samples, 2):
        row_data = [
            s['sample_code'],
            s['batch_code'],
            s['client_name'],
            s['analysis_name'],
            status_label(s['status']),
            s.get('analyst_name') or "",
            s.get('result_value') or "",
            s.get('result_notes') or "",
            s.get('result_at', "")[:10] if s.get('result_at') else "",
            s.get('expected_days') or "",
            s['created_at'][:10],
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            # Χρωματισμός κατάστασης (στήλη 5)
            if col_idx == 5:
                hex_color = STATUS_HEX.get(s['status'], "CCCCCC")
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(bold=True, color="FFFFFF")

        # Εναλλακτικό χρώμα γραμμής
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                if col_idx != 5:
                    c.fill = PatternFill("solid", fgColor="F5F7FA")

    # ── Πλάτη στηλών ──────────────────────────────────────────────────────
    col_widths = [20, 16, 20, 22, 16, 18, 14, 22, 16, 10, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # ── Sheet 2: Batches ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Batches")
    batch_headers = ["Κωδικός Batch", "Πελάτης", "Ημ. Παραλαβής", "Πλήθος Δειγμάτων"]
    for col, h in enumerate(batch_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    batches = get_all_batches(db_path)
    for row_idx, b in enumerate(batches, 2):
        ws2.cell(row=row_idx, column=1, value=b['batch_code'])
        ws2.cell(row=row_idx, column=2, value=b['client_name'])
        ws2.cell(row=row_idx, column=3, value=b['received_date'])
        ws2.cell(row=row_idx, column=4, value=b['sample_count'])

    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    wb.save(output_path)
    return len(samples)


def get_export_filename(prefix: str, extension: str) -> str:
    """Παράγει όνομα αρχείου με timestamp."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{extension}"
